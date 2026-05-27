# app.py -- Flask scheduling app v3 (monthly-first system)
import os
import calendar as cal_module
from datetime import date, timedelta, datetime
from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, flash)
from models import (db, Technician, WeekSchedule, DayAssignment, Novelty,
                    TechnicianHistory, HolidayCalendar, Config, AuditLog,
                    MonthSchedule,
                    SHIFT_T1, SHIFT_T2, SHIFT_DOMINGO, SHIFT_DESCANSO,
                    ALL_SHIFTS, SHIFT_LABELS, SHIFT_COLORS, NOVEDADES)
from scheduler import (generate_week, generate_month, apply_novelty_range,
                       get_week_stats, get_month_stats, get_month_days,
                       get_month_mondays, is_sunday_or_holiday,
                       seed_colombia_holidays, get_colombia_holidays,
                       get_alerts)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-en-produccion-2024")

DB_PATH = os.environ.get("DB_PATH", "turnos.db")
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)


# ---------------------------------------------------------------
# CONTEXT PROCESSOR
# ---------------------------------------------------------------
@app.context_processor
def inject_globals():
    return {
        "timedelta":    timedelta,
        "enumerate":    enumerate,
        "now":          datetime.utcnow(),
        "min":          min,
        "max":          max,
        "date":         date,
        "SHIFT_COLORS": SHIFT_COLORS,
        "SHIFT_LABELS": SHIFT_LABELS,
    }


# ---------------------------------------------------------------
# JINJA FILTERS
# ---------------------------------------------------------------
def date_fmt(d):
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


def weekday_es(d):
    days = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]
    return days[d.weekday()]


def day_initial(d):
    """Single letter day initial."""
    letters = ["L", "M", "M", "J", "V", "S", "D"]
    return letters[d.weekday()]


app.jinja_env.filters["date_fmt"]    = date_fmt
app.jinja_env.filters["weekday_es"]  = weekday_es
app.jinja_env.filters["day_initial"] = day_initial


# ---------------------------------------------------------------
# DB INIT + CONFIG SEED
# ---------------------------------------------------------------
def seed_config():
    defaults = [
        ("MIN_T2_DAILY", "6", "Minimo tecnicos T2 por dia"),
        ("TICKET_COUNT", "5", "Tecnicos ticket por semana"),
    ]
    changed = False
    for key, val, label in defaults:
        if Config.query.filter_by(key=key).first() is None:
            db.session.add(Config(key=key, value=val, label=label))
            changed = True
    if changed:
        db.session.commit()


with app.app_context():
    db.create_all()
    seed_config()
    # Auto-seed current year holidays on startup
    try:
        seed_colombia_holidays(date.today().year)
    except Exception:
        pass


# ---------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------
@app.route("/")
def index():
    today  = date.today()
    monday = today - timedelta(days=today.weekday())
    week   = WeekSchedule.query.filter_by(week_start=monday).first()

    # Monthly context
    ms = MonthSchedule.query.filter_by(
        year=today.year, month=today.month
    ).first()

    total_techs      = Technician.query.filter_by(is_active=True).count()
    total_months     = MonthSchedule.query.count()
    active_novelties = Novelty.query.filter(Novelty.date_end >= today).count()

    t2_today = 0
    if week:
        t2_today = DayAssignment.query.filter_by(
            week_id=week.id, date=today, shift=SHIFT_T2
        ).count()

    min_t2 = int(Config.get("MIN_T2_DAILY", "6"))

    recent_months = (MonthSchedule.query
                     .order_by(MonthSchedule.year.desc(),
                                MonthSchedule.month.desc())
                     .limit(6).all())

    recent_audit = (AuditLog.query
                    .order_by(AuditLog.timestamp.desc())
                    .limit(10).all())

    # T2 per day this week for mini chart
    chart_labels = []
    chart_t2     = []
    if week:
        for i in range(7):
            d = monday + timedelta(days=i)
            cnt = DayAssignment.query.filter_by(
                week_id=week.id, date=d, shift=SHIFT_T2
            ).count()
            chart_labels.append(weekday_es(d))
            chart_t2.append(cnt)

    # Month stats for current month
    month_stats = get_month_stats(today.year, today.month)

    # --- Dom/Festivo stats for current month ---
    month_start_d = date(today.year, today.month, 1)
    _special_das = (DayAssignment.query
                    .filter(DayAssignment.date >= month_start_d,
                            DayAssignment.date <= today,
                            DayAssignment.is_sunday_holiday == True,
                            DayAssignment.shift == SHIFT_DOMINGO)
                    .all())
    _tech_sp = {}
    for _da in _special_das:
        _tid = _da.technician_id
        if _tid not in _tech_sp:
            _t = Technician.query.get(_tid)
            _tech_sp[_tid] = {
                'name': _t.name if _t else '?',
                'sundays': 0, 'festivos': 0
            }
        if _da.date.weekday() == 6:
            _tech_sp[_tid]['sundays'] += 1
        else:
            _tech_sp[_tid]['festivos'] += 1
    tech_special_list = sorted(
        [{'id': k, **v, 'total': v['sundays'] + v['festivos']}
         for k, v in _tech_sp.items()],
        key=lambda x: x['total'], reverse=True
    )

    return render_template("index.html",
        today=today, week=week, monday=monday, ms=ms,
        total_techs=total_techs, total_months=total_months,
        active_novelties=active_novelties,
        t2_today=t2_today, min_t2=min_t2,
        recent_months=recent_months, recent_audit=recent_audit,
        chart_labels=chart_labels, chart_t2=chart_t2,
        month_stats=month_stats,
        tech_special_list=tech_special_list)


# ---------------------------------------------------------------
# MONTHLY VIEW (main operational calendar)
# ---------------------------------------------------------------
@app.route("/month")
@app.route("/month/<int:year>/<int:month>")
def month_view(year=None, month=None):
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    # Auto-seed holidays for this year
    seed_colombia_holidays(year)

    _, days_in_month = cal_module.monthrange(year, month)
    month_start = date(year, month, 1)
    month_end   = date(year, month, days_in_month)
    month_days  = get_month_days(year, month)

    # MonthSchedule record
    ms = MonthSchedule.query.filter_by(year=year, month=month).first()

    # Get all weeks overlapping this month
    mondays = get_month_mondays(year, month)

    # day_to_week: maps each date -> week slot number (1-indexed) for column filter
    day_to_week = {}
    for _wi, _mon in enumerate(mondays):
        for _i in range(7):
            _d = _mon + timedelta(days=_i)
            if _d.month == month:
                day_to_week[_d] = _wi + 1

    weeks = []
    for monday in mondays:
        wk = WeekSchedule.query.filter_by(week_start=monday).first()
        if wk:
            weeks.append(wk)

    min_t2 = int(Config.get("MIN_T2_DAILY", "6"))

    # --- Build matrix data efficiently ---
    # Single query per week > index by (tech_id, date)
    assignment_map = {}  # (tech_id, date) -> cell_dict
    for wk in weeks:
        das = DayAssignment.query.filter_by(week_id=wk.id).all()
        for da in das:
            if month_start <= da.date <= month_end:
                assignment_map[(da.technician_id, da.date)] = da.to_dict()

    # Technicians
    technicians = (Technician.query
                   .filter_by(is_active=True)
                   .order_by(Technician.name).all())

    # Daily counters
    daily_t2      = {d: 0 for d in month_days}
    daily_tickets = {d: 0 for d in month_days}
    daily_nov     = {d: 0 for d in month_days}

    rows = []
    for tech in technicians:
        cells = []
        for d in month_days:
            cell = assignment_map.get((tech.id, d))
            if not cell:
                cell = {
                    "id":               None,
                    "shift":            "--",
                    "is_ticket":        False,
                    "color":            "#9CA3AF",
                    "label":            "Sin asignar",
                    "is_sunday_holiday": is_sunday_or_holiday(d),
                    "is_manual":        False,
                    "technician_id":    tech.id,
                    "date":             d.isoformat(),
                }
            else:
                if cell["shift"] == SHIFT_T2:
                    daily_t2[d] += 1
                if cell["is_ticket"]:
                    daily_tickets[d] += 1
                if cell["shift"] in NOVEDADES:
                    daily_nov[d] += 1
            cells.append(cell)
        rows.append({"tech": tech, "cells": cells})

    # Holiday map for day headers
    from scheduler import get_colombia_holidays as _gcol
    col_holidays = _gcol(year)
    is_holiday_map = {d: d in col_holidays or d.weekday() == 6 for d in month_days}
    holiday_name_map = {d: col_holidays.get(d, "Domingo" if d.weekday() == 6 else "")
                        for d in month_days}

    # Low T2 days (working days only)
    low_t2_days = set(d for d in month_days
                      if not is_holiday_map[d] and daily_t2[d] < min_t2)

    # Prev / next month navigation
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    month_names = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                   "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    return render_template("month.html",
        year=year, month=month,
        month_start=month_start, month_end=month_end,
        month_days=month_days, days_in_month=days_in_month,
        technicians=technicians, rows=rows, ms=ms, weeks=weeks,
        daily_t2=daily_t2, daily_tickets=daily_tickets, daily_nov=daily_nov,
        is_holiday_map=is_holiday_map, holiday_name_map=holiday_name_map,
        low_t2_days=low_t2_days, min_t2=min_t2,
        all_shifts=ALL_SHIFTS, shift_labels=SHIFT_LABELS, shift_colors=SHIFT_COLORS,
        novedades_types=NOVEDADES,
        today=today,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        month_name=month_names[month - 1],
        day_to_week=day_to_week,
        mondays=mondays,
        alerts=get_alerts(year, month) if ms else [])


# ---------------------------------------------------------------
# GENERATE MONTH
# ---------------------------------------------------------------
@app.route("/generate/month", methods=["GET", "POST"])
def generate_month_view():
    today   = date.today()
    technicians = (Technician.query
                   .filter_by(is_active=True)
                   .order_by(Technician.name).all())
    min_t2      = int(Config.get("MIN_T2_DAILY", "6"))
    ticket_count = int(Config.get("TICKET_COUNT", "5"))

    if request.method == "POST":
        year_s  = request.form.get("year")
        month_s = request.form.get("month")
        force   = request.form.get("force_regenerate") == "on"
        forced_t2 = [int(x) for x in request.form.getlist("forced_t2_ids") if x]
        forced_tk = [int(x) for x in request.form.getlist("forced_ticket_ids") if x]

        try:
            yr = int(year_s)
            mo = int(month_s)
            result = generate_month(
                yr, mo,
                force_regenerate=force,
                forced_t2_ids=forced_t2 or None,
                forced_ticket_ids=forced_tk or None,
            )
            AuditLog.log("month", result["month_id"], "generate",
                         f"Mes {yr}-{mo:02d} generado. "
                         f"Semanas={result['weeks_generated']}, "
                         f"T2 total={result['total_t2']}")
            db.session.commit()

            for w in result.get("warnings", []):
                flash(f"Aviso: {w}", "warning")
            flash(
                f"Mes generado: {result['weeks_generated']} semanas nuevas, "
                f"{result['weeks_skipped']} compartidas omitidas.",
                "success"
            )
            return redirect(url_for("month_view", year=yr, month=mo))
        except ValueError as e:
            flash(str(e), "danger")

    # Preview Colombia holidays for current and next year
    preview_holidays = get_colombia_holidays(today.year)

    # Active novelties for pre-generation preview
    from datetime import date as _date
    _today = _date.today()
    active_novelties = (Novelty.query
                        .filter(Novelty.date_end >= _today)
                        .order_by(Novelty.date_start).all())

    return render_template("generate_month.html",
        today=today, technicians=technicians,
        min_t2=min_t2, ticket_count=ticket_count,
        preview_holidays=preview_holidays,
        active_novelties=active_novelties,
        novedades_types=NOVEDADES,
        SHIFT_COLORS=SHIFT_COLORS, SHIFT_LABELS=SHIFT_LABELS,
        month_names=["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                     "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"])


# ---------------------------------------------------------------
# GENERATE WEEK (kept for backward compatibility)
# ---------------------------------------------------------------
@app.route("/generate", methods=["GET", "POST"])
def generate():
    technicians = (Technician.query
                   .filter_by(is_active=True)
                   .order_by(Technician.name).all())
    min_t2       = int(Config.get("MIN_T2_DAILY", "6"))
    ticket_count = int(Config.get("TICKET_COUNT", "5"))

    if request.method == "POST":
        week_str  = request.form.get("week_start")
        force     = request.form.get("force_regenerate") == "on"
        forced_t2 = [int(x) for x in request.form.getlist("forced_t2_ids") if x]
        forced_tk = [int(x) for x in request.form.getlist("forced_ticket_ids") if x]
        try:
            ws = date.fromisoformat(week_str)
            result = generate_week(
                ws,
                force_regenerate=force,
                forced_t2_ids=forced_t2 or None,
                forced_ticket_ids=forced_tk or None,
            )
            AuditLog.log("week", result["week_id"], "generate",
                         f"Semana {ws} generada. T2={result['t2_count']}, "
                         f"Tickets={result['ticket_count']}")
            db.session.commit()
            for w in result.get("warnings", []):
                flash(f"Aviso: {w}", "warning")
            flash(f"Semana generada. T2: {result['t2_count']}, "
                  f"Tickets: {result['ticket_count']}", "success")
            return redirect(url_for("schedule", week_str=ws.isoformat()))
        except ValueError as e:
            flash(str(e), "danger")

    today  = date.today()
    monday = today - timedelta(days=today.weekday())
    return render_template("generate.html",
        technicians=technicians, today=today, monday=monday,
        min_t2=min_t2, ticket_count=ticket_count)


# ---------------------------------------------------------------
# SCHEDULE (weekly grid)
# ---------------------------------------------------------------
@app.route("/schedule")
@app.route("/schedule/<week_str>")
def schedule(week_str=None):
    today = date.today()
    if week_str:
        try:
            week_start = date.fromisoformat(week_str)
        except ValueError:
            week_start = today - timedelta(days=today.weekday())
    else:
        week_start = today - timedelta(days=today.weekday())

    week      = WeekSchedule.query.filter_by(week_start=week_start).first()
    all_weeks = WeekSchedule.query.order_by(WeekSchedule.week_start.desc()).all()

    data = None
    if week:
        week_dates = [week_start + timedelta(days=i) for i in range(7)]
        technicians = (Technician.query
                       .filter_by(is_active=True)
                       .order_by(Technician.name).all())

        daily_t2      = [0] * 7
        daily_tickets = [0] * 7
        rows = []

        for tech in technicians:
            cells = []
            for i, d in enumerate(week_dates):
                da = DayAssignment.query.filter_by(
                    week_id=week.id, technician_id=tech.id, date=d
                ).first()
                if da:
                    cell = da.to_dict()
                    if da.shift == SHIFT_T2:
                        daily_t2[i] += 1
                    if da.is_ticket and d.weekday() < 5:
                        daily_tickets[i] += 1
                else:
                    cell = {
                        "id": None, "shift": "--", "is_ticket": False,
                        "color": "#9CA3AF", "label": "Sin asignar",
                        "is_sunday_holiday": False, "is_manual": False,
                        "technician_id": tech.id, "date": d.isoformat(),
                    }
                cells.append(cell)
            rows.append({"tech": tech, "cells": cells})

        is_special = [is_sunday_or_holiday(d) for d in week_dates]
        min_t2_val = min(daily_t2) if daily_t2 else 0

        data = {
            "week": week, "dates": week_dates, "rows": rows,
            "daily_t2": daily_t2, "daily_tickets": daily_tickets,
            "is_special": is_special, "min_t2": min_t2_val,
        }

    prev_week   = week_start - timedelta(days=7)
    next_week   = week_start + timedelta(days=7)
    min_t2_cfg  = int(Config.get("MIN_T2_DAILY", "6"))

    return render_template("schedule.html",
        week_start=week_start, week=week, data=data,
        all_weeks=all_weeks, prev_week=prev_week, next_week=next_week,
        all_shifts=ALL_SHIFTS, shift_labels=SHIFT_LABELS,
        shift_colors=SHIFT_COLORS, min_t2_cfg=min_t2_cfg)


# ---------------------------------------------------------------
# NOVELTIES
# ---------------------------------------------------------------
@app.route("/novelties", methods=["GET", "POST"])
def novelties():
    today  = date.today()
    monday = today - timedelta(days=today.weekday())

    week_str = request.args.get("week_start")
    try:
        week_start = date.fromisoformat(week_str) if week_str else monday
    except ValueError:
        week_start = monday
    week_end = week_start + timedelta(days=6)

    technicians = (Technician.query
                   .filter_by(is_active=True)
                   .order_by(Technician.name).all())

    if request.method == "POST":
        tech_id      = int(request.form["technician_id"])
        date_start_s = request.form.get("date_start") or request.form.get("date")
        date_end_s   = request.form.get("date_end") or date_start_s
        nov_type     = request.form["novelty_type"]
        notes        = request.form.get("notes", "")

        try:
            d_start = date.fromisoformat(date_start_s)
            d_end   = date.fromisoformat(date_end_s)
            if d_end < d_start:
                d_end = d_start
            result = apply_novelty_range(tech_id, d_start, d_end, nov_type, notes)
            tech = Technician.query.get(tech_id)
            AuditLog.log("novelty", tech_id, "create",
                         f"{tech.name}: {nov_type} {d_start} al {d_end}")
            db.session.commit()
            for w in result.get("warnings", []):
                flash(f"Aviso: {w}", "warning")
            flash("Novedad registrada.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        return redirect(url_for("novelties", week_start=week_start.isoformat()))

    week_novelties = (Novelty.query
                      .filter(Novelty.date_start <= week_end,
                               Novelty.date_end   >= week_start)
                      .order_by(Novelty.date_start).all())

    active_novelties = (Novelty.query
                        .filter(Novelty.date_end >= today)
                        .order_by(Novelty.date_start).all())

    return render_template("novelties.html",
        today=today, week_start=week_start, week_end=week_end,
        technicians=technicians,
        week_novelties=week_novelties,
        active_novelties=active_novelties,
        novedades_types=NOVEDADES,
        SHIFT_COLORS=SHIFT_COLORS, SHIFT_LABELS=SHIFT_LABELS)


@app.route("/novelties/<int:nov_id>/delete", methods=["POST"])
def delete_novelty(nov_id):
    nov = Novelty.query.get_or_404(nov_id)
    tech_name = nov.technician.name if nov.technician else "?"
    AuditLog.log("novelty", nov_id, "delete",
                 f"Eliminada {nov.novelty_type} de {tech_name}")
    db.session.delete(nov)
    db.session.commit()
    flash("Novedad eliminada.", "success")
    return redirect(request.referrer or url_for("novelties"))


# ---------------------------------------------------------------
# API -- inline cell editing
# ---------------------------------------------------------------
@app.route("/api/assignment/<int:da_id>/shift", methods=["PATCH"])
def api_update_shift(da_id):
    da   = DayAssignment.query.get_or_404(da_id)
    week = WeekSchedule.query.get(da.week_id)

    # Check month lock
    if week:
        d = da.date
        ms = MonthSchedule.query.filter_by(year=d.year, month=d.month).first()
        if ms and ms.is_locked:
            return jsonify({"error": "Mes bloqueado"}), 403
    if week and week.is_locked:
        return jsonify({"error": "Semana bloqueada"}), 403

    body      = request.get_json(force=True)
    new_shift = body.get("shift", "").upper()
    if not new_shift or new_shift not in ALL_SHIFTS:
        return jsonify({"error": f"Turno invalido: {new_shift}"}), 400

    old_shift  = da.shift
    da.shift   = new_shift
    da.is_manual = True
    da.override_reason = body.get("reason", "Edicion manual")
    da.modified_at     = datetime.utcnow()
    if new_shift not in (SHIFT_T1,):
        da.is_ticket = False

    AuditLog.log("assignment", da_id, "update",
                 f"Turno: {old_shift} -> {new_shift}",
                 field="shift", old=old_shift, new=new_shift)
    db.session.commit()
    return jsonify(da.to_dict())


@app.route("/api/assignment/<int:da_id>/ticket", methods=["PATCH"])
def api_toggle_ticket(da_id):
    da   = DayAssignment.query.get_or_404(da_id)
    week = WeekSchedule.query.get(da.week_id)
    if week and week.is_locked:
        return jsonify({"error": "Semana bloqueada"}), 403

    da.is_ticket   = not da.is_ticket
    da.is_manual   = True
    da.modified_at = datetime.utcnow()
    AuditLog.log("assignment", da_id, "update",
                 f"Ticket: {da.is_ticket}", field="is_ticket")
    db.session.commit()
    return jsonify(da.to_dict())


# API: get daily coverage (for dynamic metric updates)
@app.route("/api/month/<int:year>/<int:month>/coverage")
def api_month_coverage(year, month):
    stats = get_month_stats(year, month)
    # Convert date keys to string
    if stats.get("t2_by_day"):
        stats["t2_by_day"] = {k.isoformat(): v
                               for k, v in stats["t2_by_day"].items()}
        stats["low_t2_days"] = [d.isoformat() for d in stats.get("low_t2_days", [])]
    return jsonify(stats)


# API: operational alerts
@app.route("/api/alerts/<int:year>/<int:month>")
def api_month_alerts(year, month):
    result = get_alerts(year, month)
    return jsonify(result)


# ---------------------------------------------------------------
# LOCK / UNLOCK MONTH
# ---------------------------------------------------------------
@app.route("/month/<int:year>/<int:month>/lock", methods=["POST"])
def lock_month(year, month):
    ms = MonthSchedule.query.filter_by(year=year, month=month).first()
    if not ms:
        flash("Mes no generado.", "danger")
        return redirect(url_for("month_view", year=year, month=month))
    ms.is_locked = True
    AuditLog.log("month", ms.id, "lock", f"Mes {year}-{month:02d} bloqueado")
    db.session.commit()
    flash("Mes bloqueado.", "info")
    return redirect(url_for("month_view", year=year, month=month))


@app.route("/month/<int:year>/<int:month>/unlock", methods=["POST"])
def unlock_month(year, month):
    ms = MonthSchedule.query.filter_by(year=year, month=month).first()
    if not ms:
        flash("Mes no generado.", "danger")
        return redirect(url_for("month_view", year=year, month=month))
    ms.is_locked = False
    AuditLog.log("month", ms.id, "unlock", f"Mes {year}-{month:02d} desbloqueado")
    db.session.commit()
    flash("Mes desbloqueado.", "info")
    return redirect(url_for("month_view", year=year, month=month))


# ---------------------------------------------------------------
# LOCK / UNLOCK WEEK
# ---------------------------------------------------------------
@app.route("/week/<int:week_id>/lock", methods=["POST"])
def lock_week(week_id):
    week = WeekSchedule.query.get_or_404(week_id)
    week.is_locked = True
    AuditLog.log("week", week_id, "lock", f"Semana {week.week_start} bloqueada")
    db.session.commit()
    flash("Semana bloqueada.", "info")
    return redirect(url_for("schedule", week_str=week.week_start.isoformat()))


@app.route("/week/<int:week_id>/unlock", methods=["POST"])
def unlock_week(week_id):
    week = WeekSchedule.query.get_or_404(week_id)
    week.is_locked = False
    AuditLog.log("week", week_id, "unlock", f"Semana {week.week_start} desbloqueada")
    db.session.commit()
    flash("Semana desbloqueada.", "info")
    return redirect(url_for("schedule", week_str=week.week_start.isoformat()))


# ---------------------------------------------------------------
# TECHNICIANS
# ---------------------------------------------------------------
@app.route("/technicians")
def technicians():
    all_techs = (Technician.query
                 .order_by(Technician.is_active.desc(), Technician.name).all())
    return render_template("technicians.html", technicians=all_techs,
                           all_shifts=[SHIFT_T1, SHIFT_T2])


@app.route("/technicians/add", methods=["GET", "POST"])
def add_technician():
    if request.method == "GET":
        return redirect(url_for("technicians"))
    name        = request.form.get("name", "").strip()
    code        = request.form.get("code", "").strip() or None
    supervisor  = request.form.get("supervisor", "").strip() or None
    fixed_shift = request.form.get("fixed_shift") or None
    tickets_only = request.form.get("tickets_only") == "on"
    no_sundays   = request.form.get("no_sundays") == "on"
    if not name:
        flash("El nombre es requerido.", "danger")
        return redirect(url_for("technicians"))
    tech = Technician(name=name, code=code, supervisor=supervisor,
                      fixed_shift=fixed_shift,
                      tickets_only=tickets_only,
                      no_sundays=no_sundays)
    db.session.add(tech)
    AuditLog.log("technician", None, "create", f"Tecnico creado: {name}")
    db.session.commit()
    flash(f"Tecnico '{name}' agregado.", "success")
    return redirect(url_for("technicians"))


@app.route("/technicians/<int:tech_id>/toggle", methods=["POST"])
def toggle_technician(tech_id):
    tech = Technician.query.get_or_404(tech_id)
    tech.is_active = not tech.is_active
    action = "activated" if tech.is_active else "deactivated"
    AuditLog.log("technician", tech_id, action,
                 f"{tech.name}: activo={tech.is_active}")
    db.session.commit()
    flash(f"{tech.name}: {'activo' if tech.is_active else 'inactivo'}.", "info")
    return redirect(url_for("technicians"))


@app.route("/technicians/<int:tech_id>/edit", methods=["POST"])
def edit_technician(tech_id):
    tech = Technician.query.get_or_404(tech_id)
    tech.name        = request.form.get("name", tech.name).strip()
    tech.supervisor  = request.form.get("supervisor", "").strip() or None
    tech.fixed_shift = request.form.get("fixed_shift") or None
    tech.tickets_only = request.form.get("tickets_only") == "on"
    tech.no_sundays   = request.form.get("no_sundays") == "on"
    AuditLog.log("technician", tech_id, "update", f"Editado: {tech.name}")
    db.session.commit()
    flash(f"Tecnico '{tech.name}' actualizado.", "success")
    return redirect(url_for("technicians"))


# ---------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------
@app.route("/config", methods=["GET", "POST"])
def config_view():
    if request.method == "POST":
        try:
            min_t2_val = request.form.get("min_t2")
            tk_val     = request.form.get("ticket_count")
            if min_t2_val:
                Config.set("MIN_T2_DAILY", int(min_t2_val),
                           "Minimo tecnicos T2 por dia")
                AuditLog.log("config", None, "update",
                             f"MIN_T2_DAILY={min_t2_val}")
            if tk_val:
                Config.set("TICKET_COUNT", int(tk_val),
                           "Tecnicos ticket por semana")
                AuditLog.log("config", None, "update",
                             f"TICKET_COUNT={tk_val}")
            db.session.commit()
            flash("Configuracion guardada.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        return redirect(url_for("config_view"))

    configs = Config.query.all()
    return render_template("config.html", configs=configs)


# ---------------------------------------------------------------
# HOLIDAYS
# ---------------------------------------------------------------
@app.route("/holidays", methods=["GET", "POST"])
def holidays():
    if request.method == "POST":
        action = request.form.get("action", "add")
        if action == "seed":
            year = int(request.form.get("seed_year", date.today().year))
            added = seed_colombia_holidays(year)
            flash(f"Festivos Colombia {year}: {added} nuevos agregados.", "success")
            return redirect(url_for("holidays"))

        date_str = request.form.get("date")
        name     = request.form.get("name", "Festivo").strip()
        try:
            d = date.fromisoformat(date_str)
            if not HolidayCalendar.query.filter_by(date=d).first():
                db.session.add(HolidayCalendar(date=d, name=name))
                db.session.commit()
                flash(f"Festivo agregado: {d} - {name}", "success")
            else:
                flash("Esa fecha ya existe.", "warning")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        return redirect(url_for("holidays"))

    all_holidays = (HolidayCalendar.query
                    .order_by(HolidayCalendar.date).all())
    today = date.today()
    return render_template("holidays.html",
                           holidays=all_holidays, today=today)


@app.route("/holidays/<int:hol_id>/delete", methods=["POST"])
def delete_holiday(hol_id):
    h = HolidayCalendar.query.get_or_404(hol_id)
    db.session.delete(h)
    db.session.commit()
    flash("Festivo eliminado.", "success")
    return redirect(url_for("holidays"))


# ---------------------------------------------------------------
# ADMIN
# ---------------------------------------------------------------
@app.route("/admin")
def admin():
    total_techs      = Technician.query.filter_by(is_active=True).count()
    inactive_techs   = Technician.query.filter_by(is_active=False).count()
    total_months     = MonthSchedule.query.count()
    locked_months    = MonthSchedule.query.filter_by(is_locked=True).count()
    total_novelties  = Novelty.query.count()
    total_holidays   = HolidayCalendar.query.count()
    configs          = Config.query.all()
    recent_audit     = (AuditLog.query
                        .order_by(AuditLog.timestamp.desc())
                        .limit(20).all())
    return render_template("admin.html",
        total_techs=total_techs, inactive_techs=inactive_techs,
        total_months=total_months, locked_months=locked_months,
        total_novelties=total_novelties, total_holidays=total_holidays,
        configs=configs, recent_audit=recent_audit)


# ---------------------------------------------------------------
# AUDIT LOG
# ---------------------------------------------------------------
@app.route("/audit")
def audit():
    page    = request.args.get("page", 1, type=int)
    per_page = 50
    logs = (AuditLog.query
            .order_by(AuditLog.timestamp.desc())
            .paginate(page=page, per_page=per_page, error_out=False))
    return render_template("audit.html", logs=logs)


# ---------------------------------------------------------------
# HISTORY / EQUITY
# ---------------------------------------------------------------
@app.route("/history")
def history():
    technicians = (Technician.query
                   .filter_by(is_active=True)
                   .order_by(Technician.name).all())
    hist_data = []
    for tech in technicians:
        rows = (TechnicianHistory.query
                .filter_by(technician_id=tech.id)
                .order_by(TechnicianHistory.week_start.desc())
                .limit(8).all())
        latest = rows[0] if rows else None
        hist_data.append({
            "tech":         tech,
            "total_t2":     latest.total_t2_weeks     if latest else 0,
            "total_tickets": latest.total_ticket_weeks if latest else 0,
            "total_sundays": latest.total_sundays      if latest else 0,
            "last_week":    latest.week_start          if latest else None,
            "rows":         rows,
        })
    hist_data.sort(key=lambda x: x["total_t2"], reverse=True)
    return render_template("history.html", hist_data=hist_data)


if __name__ == "__main__":
    app.run(debug=True, port=5000)


# ---------------------------------------------------------------
# BULK TECHNICIAN IMPORT
# ---------------------------------------------------------------

def _parse_yes(val):
    """Convert SI/YES/1/TRUE -> True, else False."""
    return str(val).strip().upper() in ("SI", "YES", "1", "TRUE", "S", "Y")


def _apply_tech_row(row_data):
    """
    Create or update a Technician from a dict.
    Returns ("created"|"updated"|"error", message)
    """
    name = str(row_data.get("nombre", "")).strip()
    if not name:
        return ("error", "Nombre vacio")
    if len(name) < 2:
        return ("error", f"Nombre muy corto: '{name}'")

    code       = str(row_data.get("codigo", "")).strip() or None
    supervisor = str(row_data.get("supervisor", "")).strip() or None
    raw_shift  = str(row_data.get("turno_fijo", "")).strip().upper()
    fixed_shift = raw_shift if raw_shift in ("T1", "T2") else None
    tickets_only = _parse_yes(row_data.get("solo_tickets", ""))
    no_sundays   = _parse_yes(row_data.get("sin_dominicales", ""))

    # Check by code first, then by name
    tech = None
    if code:
        tech = Technician.query.filter_by(code=code).first()
    if not tech:
        tech = Technician.query.filter_by(name=name).first()

    if tech:
        tech.name         = name
        tech.supervisor   = supervisor or tech.supervisor
        tech.fixed_shift  = fixed_shift or tech.fixed_shift
        tech.tickets_only = tickets_only
        tech.no_sundays   = no_sundays
        if code:
            tech.code = code
        return ("updated", name)
    else:
        # Check duplicate code
        if code and Technician.query.filter_by(code=code).first():
            return ("error", f"Codigo duplicado: {code}")
        tech = Technician(
            name=name, code=code, supervisor=supervisor,
            fixed_shift=fixed_shift,
            tickets_only=tickets_only,
            no_sundays=no_sundays,
        )
        db.session.add(tech)
        return ("created", name)


@app.route("/technicians/bulk")
def bulk_import_view():
    return render_template("bulk_import.html")


@app.route("/technicians/bulk/paste", methods=["POST"])
def bulk_paste():
    """Handle paste-list import: one name per line."""
    text = request.form.get("names_text", "")
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    if not lines:
        flash("No se recibieron datos.", "warning")
        return redirect(url_for("bulk_import_view"))

    created = []
    updated = []
    errors  = []

    for raw in lines:
        # Each line may be: "Juan Perez" or "Juan Perez, TEC001, Supervisor"
        parts = [p.strip() for p in raw.split(",")]
        row = {
            "nombre":     parts[0] if len(parts) > 0 else "",
            "codigo":     parts[1] if len(parts) > 1 else "",
            "supervisor": parts[2] if len(parts) > 2 else "",
            "turno_fijo": parts[3] if len(parts) > 3 else "",
            "solo_tickets": parts[4] if len(parts) > 4 else "",
            "sin_dominicales": parts[5] if len(parts) > 5 else "",
        }
        status, msg = _apply_tech_row(row)
        if status == "created":
            created.append(msg)
        elif status == "updated":
            updated.append(msg)
        else:
            errors.append(msg)

    db.session.commit()
    AuditLog.log("technician", None, "bulk_paste",
                 f"Paste: {len(created)} creados, {len(updated)} actualizados, {len(errors)} errores")
    db.session.commit()

    return render_template("bulk_result.html",
        method="Lista pegada",
        created=created, updated=updated, errors=errors)


@app.route("/technicians/bulk/upload", methods=["POST"])
def bulk_upload():
    """Handle Excel (.xlsx) or CSV file upload."""
    f = request.files.get("file")
    if not f or not f.filename:
        flash("No se selecciono archivo.", "warning")
        return redirect(url_for("bulk_import_view"))

    filename = f.filename.lower()
    rows_data = []
    parse_error = None

    try:
        if filename.endswith(".csv"):
            import csv, io
            content = f.read().decode("utf-8-sig")
            reader  = csv.DictReader(io.StringIO(content))
            # Normalize headers
            for row in reader:
                norm = {k.lower().strip().replace(" ", "_"): v for k, v in row.items()}
                rows_data.append(norm)

        elif filename.endswith(".xlsx"):
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).lower().strip().replace(" ", "_")
                               if c else f"col{j}" for j, c in enumerate(row)]
                else:
                    if not any(row):
                        continue
                    rows_data.append({headers[j]: (str(row[j]).strip() if row[j] is not None else "")
                                       for j in range(len(headers))})
        else:
            flash("Solo se aceptan archivos .xlsx o .csv", "danger")
            return redirect(url_for("bulk_import_view"))

    except Exception as e:
        parse_error = str(e)

    if parse_error:
        flash(f"Error al leer el archivo: {parse_error}", "danger")
        return redirect(url_for("bulk_import_view"))

    if not rows_data:
        flash("El archivo esta vacio o no tiene datos validos.", "warning")
        return redirect(url_for("bulk_import_view"))

    created = []
    updated = []
    errors  = []

    # Map flexible column names
    COL_MAP = {
        "nombre": ["nombre", "name", "tecnico", "technician", "nombres"],
        "codigo": ["codigo", "code", "cod", "id"],
        "supervisor": ["supervisor", "sup", "jefe"],
        "turno_fijo": ["turno_fijo", "turno", "fixed_shift", "shift"],
        "solo_tickets": ["solo_tickets", "tickets_only", "tickets", "ticket"],
        "sin_dominicales": ["sin_dominicales", "no_sundays", "sin_domingo"],
    }

    def get_col(row, key):
        for alias in COL_MAP.get(key, [key]):
            if alias in row:
                return row[alias]
        return ""

    for i, raw_row in enumerate(rows_data, start=2):
        row = {k: get_col(raw_row, k) for k in COL_MAP}
        status, msg = _apply_tech_row(row)
        if status == "created":
            created.append(msg)
        elif status == "updated":
            updated.append(msg)
        else:
            errors.append(f"Fila {i}: {msg}")

    db.session.commit()
    AuditLog.log("technician", None, "bulk_upload",
                 f"Upload: {len(created)} creados, {len(updated)} actualizados, {len(errors)} errores")
    db.session.commit()

    return render_template("bulk_result.html",
        method=f"Archivo: {f.filename}",
        created=created, updated=updated, errors=errors)


@app.route("/technicians/template")
def technician_template():
    """Download Excel template for bulk import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tecnicos"

    headers = [
        "nombre", "codigo", "supervisor",
        "turno_fijo", "solo_tickets", "sin_dominicales"
    ]
    labels = [
        "Nombre *", "Codigo", "Supervisor",
        "Turno Fijo (T1/T2)", "Solo Tickets (SI/NO)", "Sin Dominicales (SI/NO)"
    ]
    examples = [
        ["Juan Perez",  "TEC001", "Carlos Lopez",  "T2", "NO", "NO"],
        ["Maria Ruiz",  "TEC002", "Carlos Lopez",  "T1", "SI", "NO"],
        ["Pedro Silva", "TEC003", "Ana Gomez",     "",   "NO", "SI"],
        ["Laura Torres","TEC004", "Ana Gomez",     "T1", "NO", "NO"],
    ]

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    note_fill   = PatternFill("solid", fgColor="DBEAFE")
    note_font   = Font(color="1E40AF", italic=True, size=9)

    # Header row
    for col, (h, lbl) in enumerate(zip(headers, labels), start=1):
        cell = ws.cell(row=1, column=col, value=lbl)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Note row
    note = ws.cell(row=2, column=1,
        value="INSTRUCCIONES: Llena desde la fila 3. * = obligatorio. Turno: T1 o T2 (dejar vacio = rotacion auto).")
    note.fill = note_fill
    note.font = note_font
    ws.merge_cells("A2:F2")

    # Example rows
    for r, ex in enumerate(examples, start=3):
        for c, val in enumerate(ex, start=1):
            ws.cell(row=r, column=c, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="plantilla_tecnicos.xlsx",
    )
